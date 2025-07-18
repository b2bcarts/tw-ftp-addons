from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError
import ftplib
import tempfile
import os
import pandas as pd
import openpyxl
import logging
import socket
import paramiko
from datetime import datetime
import json
import subprocess

_logger = logging.getLogger(__name__)

class FtpService(models.Model):
    _name = 'ftp.service'
    _description = 'FTP Service Operations'

    def _get_ftp_connection(self, config):
        """Establish FTP/SFTP connection"""
        ftp = None
        ssh = None
        try:
            _logger.info(f"=== CONNECTION ATTEMPT ===")
            _logger.info(f"Config: {config.name}")
            _logger.info(f"Host: {config.host}")
            _logger.info(f"Port: {config.port}")
            _logger.info(f"Username: {config.username}")
            _logger.info(f"Connection Type: {config.connection_type}")
            _logger.info(f"Download Path: {config.download_path}")
            _logger.info(f"Processed Path: {config.processed_path}")
            
            if config.connection_type in ['sftp', 'scp']:
                _logger.info(f"Creating {config.connection_type.upper()} connection")
                ssh = paramiko.SSHClient()
                ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                
                _logger.info(f"Attempting SSH connection to {config.host}:{config.port}")
                ssh.connect(
                    hostname=config.host,
                    port=config.port,
                    username=config.username,
                    password=config.password,
                    timeout=30
                )
                _logger.info(f"✓ SSH connection established to {config.host}:{config.port}")
                
                if config.connection_type == 'sftp':
                    ftp = ssh.open_sftp()
                    _logger.info(f"✓ SFTP session opened for user {config.username}")
                    
                    # Test basic SFTP commands
                    try:
                        _logger.info("Testing SFTP pwd command")
                        current_dir = ftp.getcwd() or '/'
                        _logger.info(f"✓ Current SFTP directory: {current_dir}")
                    except Exception as pwd_error:
                        _logger.warning(f"SFTP pwd command failed: {str(pwd_error)}")
                    
                    try:
                        _logger.info("Testing SFTP dir listing")
                        file_list = ftp.listdir('.')
                        _logger.info(f"✓ SFTP dir listing successful: {len(file_list)} items")
                    except Exception as list_error:
                        _logger.warning(f"SFTP listdir command failed: {str(list_error)}")
                    
                    _logger.info("✓ Connected to SFTP successfully")
                    return {'connection': ftp, 'ssh': ssh, 'type': 'sftp'}
                
                elif config.connection_type == 'scp':
                    _logger.info(f"✓ SCP session ready for user {config.username}")
                    
                    # Test basic SCP commands via SSH
                    try:
                        _logger.info("Testing SCP pwd command")
                        stdin, stdout, stderr = ssh.exec_command('pwd')
                        current_dir = stdout.read().decode().strip()
                        _logger.info(f"✓ Current SCP directory: {current_dir}")
                    except Exception as pwd_error:
                        _logger.warning(f"SCP pwd command failed: {str(pwd_error)}")
                    
                    try:
                        _logger.info("Testing SCP dir listing")
                        stdin, stdout, stderr = ssh.exec_command(f'ls -la {config.download_path}')
                        result = stdout.read().decode().strip()
                        if result:
                            file_count = len(result.split('\n')) - 1  # Exclude header line
                            _logger.info(f"✓ SCP dir listing successful: {file_count} items")
                        else:
                            _logger.info("✓ SCP dir listing successful: empty directory")
                    except Exception as list_error:
                        _logger.warning(f"SCP dir listing command failed: {str(list_error)}")
                    
                    _logger.info("✓ Connected to SCP successfully")
                    return {'connection': None, 'ssh': ssh, 'type': 'scp'}
            
            else:
                # FTP/FTPS connection
                if config.connection_type == 'ftps' or config.use_tls:
                    _logger.info("Creating FTP_TLS instance")
                    ftp = ftplib.FTP_TLS()
                else:
                    _logger.info("Creating FTP instance")
                    ftp = ftplib.FTP()
                
                _logger.info(f"Attempting to connect to {config.host}:{config.port}")
                ftp.connect(config.host, config.port)
                _logger.info(f"✓ FTP connection established to {config.host}:{config.port}")
                
                _logger.info(f"Attempting login with user: {config.username}")
                ftp.login(config.username, config.password)
                _logger.info(f"✓ FTP login successful for user {config.username}")
                
                if config.connection_type == 'ftps' or config.use_tls:
                    _logger.info("Setting up TLS protection")
                    ftp.prot_p()
                    _logger.info("✓ FTP TLS protection enabled")
                
                # Test basic FTP commands
                try:
                    _logger.info("Testing FTP pwd command")
                    current_dir = ftp.pwd()
                    _logger.info(f"✓ Current FTP directory: {current_dir}")
                except Exception as pwd_error:
                    _logger.warning(f"FTP pwd command failed: {str(pwd_error)}")
                
                try:
                    _logger.info("Testing FTP dir listing")
                    ftp.retrlines('LIST', lambda x: _logger.info(f"FTP LIST: {x}"))
                except Exception as list_error:
                    _logger.warning(f"FTP LIST command failed: {str(list_error)}")
                
                _logger.info("✓ Connected to FTP successfully")
                return {'connection': ftp, 'ssh': None, 'type': 'ftp'}
            
        except ftplib.error_perm as perm_error:
            error_msg = f"FTP Permission Error: {str(perm_error)}"
            _logger.error(f"✗ {error_msg}")
            _logger.error("This usually means wrong username/password or insufficient permissions")
            if ftp:
                try:
                    ftp.quit()
                except:
                    pass
            raise UserError(error_msg)
            
        except ftplib.error_temp as temp_error:
            error_msg = f"FTP Temporary Error: {str(temp_error)}"
            _logger.error(f"✗ {error_msg}")
            _logger.error("This is usually a temporary server issue")
            if ftp:
                try:
                    ftp.quit()
                except:
                    pass
            raise UserError(error_msg)
            
        except socket.gaierror as dns_error:
            error_msg = f"DNS Resolution Error: {str(dns_error)}"
            _logger.error(f"✗ {error_msg}")
            _logger.error(f"Cannot resolve hostname: {config.host}")
            _logger.error("Check if the hostname is correct and accessible")
            if ftp:
                try:
                    ftp.quit()
                except:
                    pass
            raise UserError(error_msg)
            
        except socket.timeout as timeout_error:
            error_msg = f"Connection Timeout: {str(timeout_error)}"
            _logger.error(f"✗ {error_msg}")
            _logger.error(f"Connection to {config.host}:{config.port} timed out")
            _logger.error("Check if the server is running and port is correct")
            if ftp:
                try:
                    ftp.quit()
                except:
                    pass
            raise UserError(error_msg)
            
        except ConnectionRefusedError as conn_error:
            error_msg = f"Connection Refused: {str(conn_error)}"
            _logger.error(f"✗ {error_msg}")
            _logger.error(f"Server {config.host}:{config.port} refused the connection")
            _logger.error("Check if FTP service is running on the target server")
            if ftp:
                try:
                    ftp.quit()
                except:
                    pass
            raise UserError(error_msg)
            
        except Exception as e:
            error_msg = f"Unexpected FTP Error: {str(e)} (Type: {type(e).__name__})"
            _logger.error(f"✗ {error_msg}")
            _logger.error(f"Full error details: {repr(e)}")
            if ftp:
                try:
                    ftp.quit()
                except:
                    pass
            raise UserError(error_msg)
    
    def _download_file(self, connection_info, remote_filename, local_path, config):
        """Download file from FTP/SFTP/SCP to local path"""
        try:
            _logger.info(f"Starting download of file: {remote_filename}")
            
            if connection_info['type'] == 'scp':
                # Use SCP command via SSH
                ssh = connection_info['ssh']
                remote_path = f"{config.download_path}/{remote_filename}".replace('//', '/')
                
                # Use SCP via paramiko
                scp = ssh.open_sftp()
                scp.get(remote_path, local_path)
                scp.close()
                _logger.info(f"Successfully downloaded file via SCP: {remote_filename}")
                return True
                
            elif connection_info['type'] == 'sftp':
                # Use SFTP
                sftp = connection_info['connection']
                remote_path = f"{config.download_path}/{remote_filename}".replace('//', '/')
                sftp.get(remote_path, local_path)
                _logger.info(f"Successfully downloaded file via SFTP: {remote_filename}")
                return True
                
            else:
                # Use FTP
                ftp = connection_info['connection']
                with open(local_path, 'wb') as local_file:
                    ftp.retrbinary(f'RETR {remote_filename}', local_file.write)
                _logger.info(f"Successfully downloaded file via FTP: {remote_filename}")
                return True
                
        except Exception as e:
            _logger.error(f"Failed to download {remote_filename}: {str(e)}")
            return False
    
    def _process_excel_file(self, file_path):
        """Process Excel file and return content as dictionary with first row as keys"""
        try:
            _logger.info(f"Processing Excel file: {file_path}")
            # Try to read with openpyxl first (better for .xlsx)
            if file_path.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                content = {}
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))
                    
                    if not rows:
                        _logger.warning(f"Sheet {sheet_name} is empty")
                        content[sheet_name] = []
                        continue
                    
                    # Get headers from first row
                    headers = []
                    for cell in rows[0]:
                        if cell is None:
                            headers.append("")
                        else:
                            headers.append(str(cell))
                    
                    # Process data rows (skip first row which contains headers)
                    sheet_data = []
                    for row in rows[1:]:
                        row_dict = {}
                        for i, cell in enumerate(row):
                            # Use header as key, or fallback to column index
                            key = headers[i] if i < len(headers) else f"column_{i}"
                            
                            # Convert cell value
                            if cell is None:
                                value = ""
                            elif isinstance(cell, datetime):
                                value = cell.isoformat()
                            else:
                                value = str(cell)
                            
                            row_dict[key] = value
                        
                        # Only add non-empty rows
                        if any(value.strip() for value in row_dict.values() if isinstance(value, str)):
                            sheet_data.append(row_dict)
                    
                    content[sheet_name] = sheet_data
                    _logger.info(f"Sheet {sheet_name}: {len(sheet_data)} rows processed with headers: {headers}")
                
                workbook.close()
                _logger.info(f"Successfully processed Excel file with {len(content)} sheets")
                return content
            
            # Fallback to pandas for .xls files
            else:
                excel_file = pd.ExcelFile(file_path)
                content = {}
                
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    if df.empty:
                        _logger.warning(f"Sheet {sheet_name} is empty")
                        content[sheet_name] = []
                        continue
                    
                    # Replace NaN with empty strings
                    df = df.fillna('')
                    
                    # Convert DataFrame to list of dictionaries
                    sheet_data = []
                    for _, row in df.iterrows():
                        row_dict = {}
                        for column in df.columns:
                            value = row[column]
                            if isinstance(value, pd.Timestamp):
                                value = value.isoformat()
                            else:
                                value = str(value)
                            row_dict[str(column)] = value
                        
                        # Only add non-empty rows
                        if any(value.strip() for value in row_dict.values() if isinstance(value, str)):
                            sheet_data.append(row_dict)
                    
                    content[sheet_name] = sheet_data
                    _logger.info(f"Sheet {sheet_name}: {len(sheet_data)} rows processed with headers: {list(df.columns)}")
                
                _logger.info(f"Successfully processed Excel file with {len(content)} sheets")
                return content
                
        except Exception as e:
            _logger.error(f"Failed to process Excel file {file_path}: {str(e)}")
            raise UserError(f"Failed to process Excel file: {str(e)}")
    
    def _move_file_on_connection(self, connection_info, filename, new_path, config):
        """Move file based on connection type"""
        try:
            _logger.info(f"Moving file {filename} to {new_path}")
            old_path = f"{config.download_path}/{filename}".replace('//', '/')
            
            if connection_info['type'] == 'sftp':
                sftp = connection_info['connection']
                # Create destination directory if it doesn't exist
                dest_dir = os.path.dirname(new_path)
                if dest_dir != '/':
                    try:
                        sftp.makedirs(dest_dir)
                        _logger.info(f"Created directory: {dest_dir}")
                    except Exception as e:
                        _logger.info(f"Directory {dest_dir} already exists or cannot be created: {str(e)}")
                
                sftp.rename(old_path, new_path)
                _logger.info(f"Successfully moved file from {old_path} to {new_path}")
                return True
                
            elif connection_info['type'] in ['ftp', 'ftps']:
                ftp = connection_info['connection']
                # Create destination directory if it doesn't exist
                dest_dir = os.path.dirname(new_path)
                if dest_dir != '/':
                    try:
                        ftp.mkd(dest_dir)
                        _logger.info(f"Created directory: {dest_dir}")
                    except ftplib.error_perm as e:
                        _logger.info(f"Directory {dest_dir} already exists or cannot be created: {str(e)}")
                
                ftp.rename(filename, new_path)
                _logger.info(f"Successfully moved file from {filename} to {new_path}")
                return True
            
            else:
                _logger.warning(f"Move operation not supported for {connection_info['type']}")
                return False
                
        except Exception as e:
            _logger.error(f"Failed to move file from {filename} to {new_path}: {str(e)}")
            return False
    
    def _get_file_size(self, connection_info, filename, config):
        """Get file size from connection"""
        try:
            if connection_info['type'] == 'sftp':
                sftp = connection_info['connection']
                remote_path = f"{config.download_path}/{filename}".replace('//', '/')
                stat = sftp.stat(remote_path)
                size = stat.st_size / 1024  # Convert to KB
                _logger.info(f"File {filename} size: {size:.2f} KB")
                return size
            elif connection_info['type'] == 'scp':
                ssh = connection_info['ssh']
                remote_path = f"{config.download_path}/{filename}".replace('//', '/')
                stdin, stdout, stderr = ssh.exec_command(f'stat -c%s {remote_path}')
                result = stdout.read().decode().strip()
                if result.isdigit():
                    size = int(result) / 1024  # Convert to KB
                    _logger.info(f"File {filename} size: {size:.2f} KB")
                    return size
                else:
                    _logger.warning(f"Could not get file size for {filename}")
                    return 0.0
            else:
                # FTP
                ftp = connection_info['connection']
                size = ftp.size(filename) / 1024  # Convert to KB
                _logger.info(f"File {filename} size: {size:.2f} KB")
                return size
        except Exception as e:
            _logger.warning(f"Could not get file size for {filename}: {str(e)}")
            return 0.0
    
    def _close_connection(self, connection_info):
        """Close connection properly"""
        try:
            if connection_info['type'] in ['sftp', 'scp']:
                if connection_info['connection']:
                    connection_info['connection'].close()
                if connection_info['ssh']:
                    connection_info['ssh'].close()
            else:
                if connection_info['connection']:
                    connection_info['connection'].quit()
            _logger.info(f"Connection closed successfully")
        except Exception as e:
            _logger.warning(f"Error closing connection: {str(e)}")
    
    @api.model
    def process_ftp_files(self, config_id=None):
        """Main method to process FTP files"""
        configs = self.env['ftp.config'].browse(config_id) if config_id else self.env['ftp.config'].search([('active', '=', True)])
        
        for config in configs:
            _logger.info(f"Processing files for config: {config.name}")
            
            try:
                connection_info = self._get_ftp_connection(config)
                connection_type = connection_info['type']
                
                # Get list of Excel files based on connection type
                file_list = []
                try:
                    if connection_type in ['sftp', 'scp']:
                        if connection_type == 'sftp':
                            sftp = connection_info['connection']
                            try:
                                sftp.chdir(config.download_path)
                                _logger.info(f"Changed to directory: {config.download_path}")
                            except Exception as e:
                                _logger.warning(f"Could not change to directory {config.download_path}: {str(e)}")
                            
                            file_list = sftp.listdir('.')
                            _logger.info(f"Found {len(file_list)} files in SFTP directory")
                        
                        elif connection_type == 'scp':
                            ssh = connection_info['ssh']
                            stdin, stdout, stderr = ssh.exec_command(f'ls {config.download_path}')
                            result = stdout.read().decode().strip()
                            if result:
                                file_list = result.split('\n')
                                _logger.info(f"Found {len(file_list)} files in SCP directory")
                            else:
                                _logger.info("No files found in SCP directory")
                    
                    else:
                        # FTP/FTPS connection
                        ftp = connection_info['connection']
                        try:
                            ftp.cwd(config.download_path)
                            _logger.info(f"Changed to directory: {config.download_path}")
                        except Exception as e:
                            _logger.error(f"Failed to change to directory {config.download_path}: {str(e)}")
                            ftp.quit()
                            continue
                        
                        ftp.retrlines('NLST', file_list.append)
                        _logger.info(f"Found {len(file_list)} files in FTP directory")
                
                except Exception as e:
                    _logger.error(f"Failed to list files in directory {config.download_path}: {str(e)}")
                    self._close_connection(connection_info)
                    continue
                
                excel_files = [f for f in file_list if f.lower().endswith(('.xlsx', '.xls'))]
                _logger.info(f"Found {len(excel_files)} Excel files: {excel_files}")
                
                for filename in excel_files:
                    # Check if file already processed
                    try:
                        existing_file = self.env['ftp.file'].search([
                            ('name', '=', filename),
                            ('ftp_config_id', '=', config.id),
                            ('status', 'in', ['processed', 'moved'])
                        ], limit=1)
                        
                        if existing_file:
                            _logger.info(f"File {filename} already processed, skipping")
                            continue
                    except Exception as e:
                        _logger.error(f"Error checking existing file {filename}: {str(e)}")
                        continue
                    
                    # Download and process file
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp_file:
                        tmp_path = tmp_file.name
                    
                    try:
                        # Download file
                        if self._download_file(connection_info, filename, tmp_path, config):
                            try:
                                file_size = self._get_file_size(connection_info, filename, config)
                                
                                # Process Excel content
                                content = self._process_excel_file(tmp_path)
                                
                                # Calculate statistics
                                total_rows = sum(len(sheet_data) for sheet_data in content.values())
                                total_cols = 0
                                for sheet_data in content.values():
                                    if sheet_data:  # If sheet has data
                                        sheet_cols = len(sheet_data[0]) if sheet_data else 0
                                        total_cols = max(total_cols, sheet_cols)
                                sheet_names = list(content.keys())
                                
                                # Create file record
                                file_record = self.env['ftp.file'].create({
                                    'name': filename,
                                    'file_size': file_size,
                                    'ftp_config_id': config.id,
                                    'content_json': json.dumps(content, ensure_ascii=False, indent=2),
                                    'original_path': config.download_path + '/' + filename,
                                    'status': 'processed',
                                    'row_count': total_rows,
                                    'column_count': total_cols,
                                    'sheet_names': ', '.join(sheet_names)
                                })
                                _logger.info(f"Created file record for: {filename}")
                                
                                # Move file to processed directory (only for FTP/SFTP for now)
                                if connection_info['type'] in ['ftp', 'ftps', 'sftp']:
                                    new_path = config.processed_path + '/' + filename
                                    if self._move_file_on_connection(connection_info, filename, new_path, config):
                                        file_record.write({
                                            'moved_path': new_path,
                                            'status': 'moved'
                                        })
                                        _logger.info(f"Successfully processed and moved file: {filename}")
                                    else:
                                        _logger.warning(f"File processed but not moved: {filename}")
                                else:
                                    _logger.info(f"File processing completed for {filename} (move not supported for {connection_info['type']})")
                                    
                            except Exception as e:
                                _logger.error(f"Error processing file content for {filename}: {str(e)}")
                                raise e
                        else:
                            _logger.error(f"Failed to download file: {filename}")
                            raise Exception(f"Download failed for file: {filename}")
                        
                    except Exception as e:
                        # Create error record
                        try:
                            self.env['ftp.file'].create({
                                'name': filename,
                                'file_size': self._get_file_size(connection_info, filename, config),
                                'ftp_config_id': config.id,
                                'original_path': config.download_path + '/' + filename,
                                'status': 'error',
                                'error_message': str(e)
                            })
                            _logger.error(f"Error processing file {filename}: {str(e)}")
                        except Exception as create_error:
                            _logger.error(f"Failed to create error record for {filename}: {str(create_error)}")
                    
                    finally:
                        # Clean up temporary file
                        try:
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
                                _logger.info(f"Cleaned up temporary file: {tmp_path}")
                        except Exception as e:
                            _logger.warning(f"Failed to clean up temporary file {tmp_path}: {str(e)}")
                
                # Update last sync time and close connection
                try:
                    config.last_sync = fields.Datetime.now()
                    self._close_connection(connection_info)
                    _logger.info(f"File processing completed for config: {config.name}")
                except Exception as e:
                    _logger.error(f"Error closing connection for {config.name}: {str(e)}")
                
            except Exception as e:
                _logger.error(f"Critical error in file processing for {config.name}: {str(e)}")
                # Try to close connection if it exists
                try:
                    if 'connection_info' in locals():
                        self._close_connection(connection_info)
                except:
                    pass
                continue
    
    @api.model
    def cron_process_ftp_files(self):
        """Cron job method"""
        try:
            _logger.info("Starting scheduled FTP file processing")
            self.process_ftp_files()
            _logger.info("Scheduled FTP file processing completed")
        except Exception as e:
            _logger.error(f"Error in scheduled FTP processing: {str(e)}")
    
    def reprocess_file(self, file_id):
        """Reprocess a specific file record"""
        try:
            file_record = self.env['ftp.file'].browse(file_id)
            if file_record and file_record.moved_path:
                # This would require downloading from the moved location
                # Implementation depends on specific requirements
                _logger.info(f"Reprocessing file: {file_record.name}")
                # Add reprocessing logic here if needed
            else:
                _logger.warning(f"File record {file_id} not found or no moved path")
        except Exception as e:
            _logger.error(f"Error reprocessing file {file_id}: {str(e)}")